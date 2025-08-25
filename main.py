import datetime
import math
import os
import sys
from time import sleep
import RepkaPi.GPIO as GPIO
#import serial
from PyQt5.QtCore import QThread, pyqtSignal
import pandas as pd
import numpy as np
from PyQt5 import QtWidgets, uic
import easygui
import openpyxl
import pyqtgraph as pg
import subprocess
import time

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.position_v = None
        uic.loadUi('/root/raspberry_app/main.ui', self)

        # UI элементы
        self.date = self.findChild(QtWidgets.QDateEdit, 'date')
        self.open_button = self.findChild(QtWidgets.QPushButton, 'open_button')
        self.file_name_display = self.findChild(QtWidgets.QTextBrowser, 'file_name')
        self.open_settings = self.findChild(QtWidgets.QPushButton, 'open_settings')
        self.keyboard = self.findChild(QtWidgets.QPushButton, 'keyboard_button')
        self.sheetName = self.findChild(QtWidgets.QPlainTextEdit, 'sheet_name')
        self.calculate = self.findChild(QtWidgets.QPushButton, 'save_button')
        self.saveSheetButton = self.findChild(QtWidgets.QPushButton, 'save_button_2')
        self.time_izm = self.findChild(QtWidgets.QComboBox, 'time_izm')
        self.status = self.findChild(QtWidgets.QTextBrowser, 'status')
        self.position_v = self.findChild(QtWidgets.QTextBrowser, 'position_V')

        # График
        self.graphWidget = self.findChild(pg.PlotWidget, 'graphWidget')
        self.graphWidget.setBackground('w')
        self.graphWidget.setLabel('left', 'Сопротивление, Ом', **{'font-size': '20pt'})
        self.graphWidget.setLabel('bottom', 'Время, сек', **{'font-size': '20pt'})
        self.graphWidget.showGrid(x=True, y=True)
        self.graphWidget.getAxis('left').setPen(pg.mkPen(color='k'))
        self.graphWidget.getAxis('bottom').setPen(pg.mkPen(color='k'))
        self.graphWidget.getAxis('left').setTextPen('k')
        self.graphWidget.getAxis('bottom').setTextPen('k')
        legend = self.graphWidget.addLegend(offset=(400, 300))
        legend.labelTextColor = pg.mkColor('k')

        # GPIO
        self.position_V = 500
        #GPIO.setwarnings(False)
        GPIO.setmode(GPIO.BOARD)
        try:
            GPIO.setup(35, GPIO.IN)
            GPIO.setup(19, GPIO.IN)
            GPIO.setup(21, GPIO.IN)
            GPIO.setup(23, GPIO.IN)
        except Exception as e:
            print(f"Error setting up GPIO: {e}")

        self.position_v.setText(str(self.position_V))

        # Кнопки
        self.open_button.clicked.connect(self.showDialog)
        self.calculate.clicked.connect(self.doCalculation)
        self.saveSheetButton.clicked.connect(self.saveSheet)
        self.open_settings.clicked.connect(self.open_window_settings)

        self.input_file = None
        self.basic_flag = 0
        self.R_itog_array = []

        # Выводы
        self.R15 = self.findChild(QtWidgets.QTextBrowser, 'R15')
        self.R30 = self.findChild(QtWidgets.QTextBrowser, 'R30')
        self.R60 = self.findChild(QtWidgets.QTextBrowser, 'R60')
        self.R600 = self.findChild(QtWidgets.QTextBrowser, 'R600')
        self.Kabs = self.findChild(QtWidgets.QTextBrowser, 'Kabs')
        self.PI = self.findChild(QtWidgets.QTextBrowser, 'PI')
        self.DAR = self.findChild(QtWidgets.QTextBrowser, 'DAR')
        self.DD = self.findChild(QtWidgets.QTextBrowser, 'DD')
        self.W = self.findChild(QtWidgets.QTextBrowser, 'W')
        self.DP = self.findChild(QtWidgets.QTextBrowser, 'DP')
        self.Res = self.findChild(QtWidgets.QTextBrowser, 'Res')

        self.C = 0
        self.I = 0

        # Поток мониторинга GPIO
        self.gpio_thread = GPIOMonitorThread()
        self.gpio_thread.position_changed.connect(self.update_position_v)
        self.gpio_thread.start()

        self.button_thread = ButtonThread()
        self.button_thread.button_pressed.connect(self.on_button_pressed)
        self.button_thread.start()



        self.showFullScreen()

    def on_button_pressed(self):
        print("Кнопка нажата! Запускаем нужную функцию")
        self.start_com()

    def update_position_v(self, new_position_v):
        if self.position_V != new_position_v:
            self.position_V = new_position_v
            self.position_v.setText(str(self.position_V))
            print(f"Режим изменён на {new_position_v}В")

    def closeEvent(self, event):
        self.gpio_thread.stop()
        self.gpio_thread.wait()
        event.accept()



    def update_status(self):
        self.status.setText(self.message)

    def open_window_settings(self):
        try:
            self.second_window = SettingsWindow()
            self.second_window.show()
            self.second_window.move(0, 0)
        except Exception as e:
            print(f"Ошибка при открытии окна настроек {e}")


    def convert_amperes(self, value):
        data = {
            'Unit': ['pA', 'nA', 'µA', 'mA', 'A', 'kA', 'MA', 'GA', 'TA'],
            'Value': [1e-12, 1e-9, 1e-6, 1e-3, 1, 1e3, 1e6, 1e9, 1e12]
        }
        Current_Units = pd.DataFrame(data)

        if value == 0:
            return "0 A"

        abs_value = abs(value)
        log_value = math.log10(abs_value)

        # Находим ближайший префикс
        index = min(range(len(Current_Units['Value'])),
                    key=lambda i: abs(math.log10(Current_Units['Value'][i]) - log_value))

        unit = Current_Units.loc[index, 'Unit']
        scale = Current_Units.loc[index, 'Value']

        scaled_value = value / scale

        # Округляем до 3 значащих цифр
        rounded_value = round(scaled_value, 2 - int(math.floor(math.log10(abs(scaled_value)))))

        return f"{rounded_value} {unit}"

    def convert_farads(self, value):
        data = {
            'Unit': ['pF', 'nF', 'µF', 'mF', 'F', 'kF', 'MF', 'GF', 'TF'],
            'Value': [1e-12, 1e-9, 1e-6, 1e-3, 1, 1e3, 1e6, 1e9, 1e12]
        }
        Capacitance_Units = pd.DataFrame(data)

        if value == 0:
            return "0 F"

        abs_value = abs(value)
        log_value = math.log10(abs_value)

        # Находим ближайший префикс
        index = min(range(len(Capacitance_Units['Value'])),
                    key=lambda i: abs(math.log10(Capacitance_Units['Value'][i]) - log_value))

        unit = Capacitance_Units.loc[index, 'Unit']
        scale = Capacitance_Units.loc[index, 'Value']

        scaled_value = value / scale

        # Округляем до 3 значащих цифр
        rounded_value = round(scaled_value, 2 - int(math.floor(math.log10(abs(scaled_value)))))

        return f"{rounded_value} {unit}"

    def showDialog(self):
        self.input_file = easygui.fileopenbox()
        if self.input_file:  # Если файл был выбран
            self.file_name_display.setPlainText(self.input_file)
        else:
            self.file_name_display.setPlainText("No file selected")

    def showKeyboard(self):
        print("click button")
        subprocess.run(['onboard'])

    def doCalculation(self):
        book = openpyxl.load_workbook(filename=self.input_file)
        sheet = book['Лист1']
        U_test = sheet['J2'].value  # считываем значение тестового напряжения из ячейки U(Volt)
        DAR_test = sheet['K2'].value  # считываем параметр DAR
        PI_test = sheet['L2'].value  # считываем параметр PI
        DD_test = sheet['O2'].value  # считываем параметр DD

        Tg = 45  # Время жизни трансформатора - Константа в годах

        try:
            I = str(sheet['N2'].value).replace(',', '.')  # Заменяем запятую на точку

            data = {
                'Unit': ['p', 'n', 'µ', 'm', ' ', 'k', 'M', 'G', 'T'],
                'Value': [1e-12, 1e-9, 1e-6, 1e-3, 1, 1e3, 1e6, 1e9, 1e12]
            }
            Razmernost = pd.DataFrame(data)

            if 'E' in I or 'e' in I:
                I_test = float(I)  # Обрабатываем научную нотацию
            else:
                n = len(I)
                unit = I[n - 2]  # Предполагаем, что последний символ — пробел, а перед ним — единица измерения
                unit_index = Razmernost[Razmernost['Unit'] == unit].index
                value = Razmernost.loc[unit_index[0], 'Value'] if not unit_index.empty else 1
                I_test = float(I[:n - 3]) * value

            Cap = str(sheet['M2'].value).replace(',', '.')

            if 'E' in Cap or 'e' in Cap:
                C_test = float(Cap)
            else:
                n = len(Cap)
                unit = Cap[n - 2]
                unit_index = Razmernost[Razmernost['Unit'] == unit].index
                value = Razmernost.loc[unit_index[0], 'Value'] if not unit_index.empty else 1
                C_test = float(Cap[:n - 3]) * value

        except Exception as e:
            print(e)


        # self.time = self.findChild(QtWidgets.QSpinBox, 'time')
        Tizm = []
        Uizm = []
        R_meas = []

        row = 2
        while True:
            t_cell = sheet['R' + str(row)]
            u_cell = sheet['S' + str(row)]
            r_cell = sheet['T' + str(row)]

            if t_cell.value is None:
                break
            time_izm = int(t_cell.value)
            Tizm.append(int(t_cell.value))
            Uizm.append(int(u_cell.value))
            R_meas.append(int(r_cell.value) * 10 ** 6)

            row += 1

        I_t = np.array(Uizm) / np.array(R_meas)

        p = np.polyfit(Tizm, R_meas, 4)
        R_apr = np.polyval(p, Tizm)
        if time_izm > 100:
            I_apr = np.polyval(np.polyfit(Tizm[21:], I_t[21:], 4), Tizm)

        self.graphWidget.clear()
        self.graphWidget.plot(Tizm, R_meas, pen=pg.mkPen(color='b', width=5), name='R измеренное')
        self.graphWidget.plot(Tizm, R_apr, pen=pg.mkPen(color='k', width=5), name='R апроксимированное')

        DAR = R_apr[10] / R_apr[1]
        self.DAR.setText(str(round(DAR, 3)))
        if (time_izm // 5 + 1 < 121):
            PI = 0
            DD = 0
        else:
            PI = R_apr[118] / R_apr[10]
            DD = I_test / (Uizm[1] * C_test)

        self.PI.setText(str(round(PI, 3)))
        self.DD.setText(str(round(DD, 3)))

        if PI == 0:
            W = 0
            TPI = 0
            Res = 0
        else:
            W = 3.275 - 4.819 * math.log10(PI)
            TPI = 59.029 * PI - 56.391
            Res = (70 - Tg) * ((TPI / 3) ** 0.251 - 1)

        self.W.setText(str(round(W, 3)))
        self.Res.setText(str(math.trunc(Res)))

        Kabs = R_apr[10] / R_apr[3]
        self.Kabs.setText(str(round(Kabs, 3)))
        DP = 200 * TPI ** 0.251
        self.DP.setText(str(math.trunc(DP)))
        R15 = R_apr[2] / 10 ** 9
        self.R15.setText(str(round(R15, 3)))

        if len(R_apr) > 15:
            R30 = R_apr[4] / 10 ** 9
            self.R30.setText(str(round(R30, 3)))
            R60 = R_apr[10] / 10 ** 9
            self.R60.setText(str(round(R60, 3)))
            R600 = R_apr[118] / 10 ** 9
            self.R600.setText(str(round(R600, 3)))
        else:
            R30 = R_apr[4] / 10 ** 9
            self.R30.setText(str(round(R30, 3)))
            R60 = R_apr[10] / 10 ** 9
            self.R60.setText(str(round(R60, 3)))
            R600 = 0
            self.R600.setText(str(round(R600, 3)))

        if time_izm > 100:
            I_ut = min(I_apr)
            I_spectr = (I_apr - I_ut) * time  # особое внимание этой строчке

    def saveSheet(self):
        sheet = openpyxl.Workbook()
        sheet.create_sheet("Лист1")
        book = sheet['Лист1']
        process = subprocess.Popen(['python3', 'save_window.py'])
        # присваивание статических значений первой строки
        book['A1'].value = "Obj:Tst"
        book['B1'].value = "Description"
        book['C1'].value = "Location"
        book['D1'].value = "Date"
        book['E1'].value = "Time"
        book['F1'].value = "Function"
        book['G1'].value = "R"

        book['H1'].value = "Unit"
        book['I1'].value = "R (T° corrected)"
        book['J1'].value = "U (Volt)"
        book['J2'].value = self.position_V
        book['K1'].value = "DAR"
        book['K2'].value = self.DAR.toPlainText()
        book['L1'].value = "PI"
        book['L2'].value = self.PI.toPlainText()
        book['M1'].value = "C"
        book['M2'].value = self.C
        book['N1'].value = "I"
        book['N2'].value = self.I

        book['O1'].value = "DD"
        book["O2"].value = self.DD.toPlainText()
        book['P1'].value = "Comments"
        book['Q1'].value = "Next Test"
        book['R1'].value = "Time"
        book['S1'].value = "U (Volt)"
        book['T1'].value = "R (MOhm)"
        book['U1'].value = "R (MOhm)T° corrected:40C"

        R_izm = self.R_itog_array

        # присваивание динамических значений второй и последующих строк
        time = int(self.time_izm.currentText()) * 60
        book['D2'].value = self.date.date().isoformat()
        book['E2'].value = datetime.datetime.now().strftime('%H:%M:%S')

        length = len(R_izm)  # Замените на нужную длину

        Tizm = [5 * i for i in range(length)]

        p = np.polyfit(Tizm, R_izm, 4)
        R_apr = np.polyval(p, Tizm)

        default_position = 0
        default_time_position = 10
        # заполнение ячеек со значениями
        for i in range(2, len(R_izm) + 2):
            column = "R" + str(i)
            book[column].value = default_time_position
            column = "S" + str(i)
            book[column].value = self.position_V
            column = "T" + str(i)
            book[column].value = R_izm[default_position] // 1000000
            column = "U" + str(i)
            book[column].value = R_apr[default_position] // 1000000
            default_position = default_position + 1
            default_time_position = default_time_position + 5

        book = sheet.create_sheet('Метаданные')

        df = pd.read_csv("metadata.csv")
        book['A1'] = "Объект"
        book['A2'] = str(df.loc[0, "object"])
        book['B1'] = "Локация"
        book['B2'] = str(df.loc[0, "location"])
        book['C1'] = "Дата"
        book['C2'] = self.date.date().isoformat()
        book['D1'] = "Оператор"
        book['D2'] = str(df.loc[0, "operator"])
        book['E1'] = "Номер измерения"
        book['E2'] = str(df.loc[0, "number_measurment"])

        sheet_name = (
                str(df.loc[0, "object"]) + " " +
                str(df.loc[0, "location"]) + " " +
                str(datetime.datetime.now().strftime('%d-%m-%Y')) + " " +
                str(df.loc[0, "operator"]) + " " +
                str(df.loc[0, "number_measurment"])
        )
        # Полное имя файла
        file_name = sheet_name + ".xlsx"

        # Проверяем, существует ли файл
        if os.path.exists(file_name):
            print(f"Файл {file_name} уже существует. Добавляем суффикс.")
            counter = 1
            while os.path.exists(f"{sheet_name}_{counter}.xlsx"):
                counter += 1
            file_name = f"{sheet_name}_{counter}.xlsx"

        # Сохраняем файл
        sheet.save(file_name)
        print(f"Файл сохранен как {file_name}")
        sheet.close()
        time.sleep(2)
        process.kill()

    def start_com(self):
        try:
            process = subprocess.Popen(['python3', 'init_commands.py'])
            port_name = "COM4"
            # Открытие последовательного порта
            with serial.Serial("/dev/ttyUSB0", baudrate=9600, timeout=1) as ser:
                time.sleep(1)  # Подождите, пока порт откроется
                print(f"Serial port {port_name} open")
                time_izm = int(self.time_izm.currentText())
                self.message = "Выполняется отправка команд"
                start_commands = [
                    "40526E0D0A",
                    "4055660D0A",
                    "49640D0A",
                    "4054720D0A",
                    "45723030300D0A",
                    "45723030310D0A",
                    "45723030320D0A",
                    "40547332342E30382E31342031323A35330D0A",
                    "54720D0A"
                ]
                if self.basic_flag == 0:
                    for cmd in start_commands:
                        hex_cmd = bytes.fromhex(cmd)
                        print(f"Sending command: {cmd}")
                        ser.write(hex_cmd)
                        ser.flush()  # Убедитесь, что данные записаны в порт
                        output = ser.readline()
                        print(f"Received output: {output}")
                        time.sleep(1)  # Пауза между командами (если необходимо)
                    self.basic_flag = 1
                commands = []
                if self.position_V == 500 and time_izm == 1:
                    commands = [
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303041303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303035303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383030336330303035303030410D0A",
                        "4466666666660D0A",
                        "467332310D0A",
                        "42640D0A"
                    ]
                elif self.position_V == 1000 and time_izm == 1:
                    commands = [
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303041303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303035303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383030336330303035303030410D0A",
                        "4466666666660D0A",
                        "467332320D0A",
                        "42640D0A"
                    ]
                elif self.position_V == 2500 and time_izm == 1:
                    commands = [
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303041303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303035303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383030336330303035303030410D0A",
                        "4466666666660D0A",
                        "467332330D0A",
                        "42640D0A"
                    ]
                elif self.position_V == 500 and time_izm == 10:
                    commands = [
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032363230303035303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303035303030410D0A",
                        "4466666666660D0A",
                        "467332310D0A",
                        "42640D0A"
                    ]
                elif self.position_V == 1000 and time_izm == 10:
                    commands = [
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032363230303035303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303035303030410D0A",
                        "4466666666660D0A",
                        "467332320D0A",
                        "42640D0A"
                    ]
                elif self.position_V == 2500 and time_izm == 10:
                    commands = [
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032363230303035303030410D0A",
                        "404045723030300D0A",
                        "457730303030453033334233433033314530303343303235383032353830303035303030410D0A",
                        "4466666666660D0A",
                        "467332330D0A",
                        "42640D0A"
                    ]
                # Отправка команд
                for cmd in commands:
                    hex_cmd = bytes.fromhex(cmd)
                    print(f"Sending command: {cmd}")
                    ser.write(hex_cmd)
                    ser.flush()  # Убедитесь, что данные записаны в порт
                    output = ser.readline().decode("utf-8")
                    while "EC" in output:
                        time.sleep(1)
                        ser.write(hex_cmd)
                        time.sleep(1)
                        output = ser.readline().decode("utf-8")
                    print(f"Received output: {output}")
                    time.sleep(3)  # Пауза между командами (если необходимо)
                process.kill()
                process = subprocess.Popen(['python3', 'measurment_timer.py', str(int(self.time_izm.currentText()) * 61)])
                # Чтение данных после отправки команд
                print("Reading data from serial port...")
                time.sleep(2)  # Дайте время устройству для отправки данных
                time_array = []
                R_array = []
                for i in range(time_izm * 60 + 5):
                    time.sleep(1)
                end_output = ""
                process.kill()
                process = subprocess.Popen(['python3', 'calculation.py'])
                while len(end_output) < 50:
                    ser.write(bytes.fromhex("4044700D0A"))
                    sleep(2)
                    end_output = ser.readline()
                    print(end_output)
                sleep(1)
                time_array = []
                volt_array = []
                R_array = []
                base_index = 2
                decoded_output = end_output.decode("utf-8")
                end_array = decoded_output.split(";")
                print('Array length' + " " + str(len(end_array)))
                for i in range(0, time_izm * 60, 5):
                    volt_array.append(int(self.position_V))
                    time_array.append(i)
                    R = end_array[base_index].split("E")
                    if i == time_izm * 60:
                        R[0] = R[0][1:]
                        itogR = float(R[0]) * 10 ** int(R[1][:-4])
                    else:
                        R[0] = R[0][1:]
                        itogR = float(R[0]) * 10 ** int(R[1])
                    R_array.append(itogR)
                    base_index += 2
                self.R_itog_array = R_array

                print("считывание финального измерения")
                ser.write(bytes.fromhex("44670D0A"))
                sleep(2)
                output = ser.readline()
                print(f"Received output: {output}")
                output = output.decode("utf-8")
                result_array = output.split(";")
                R = result_array[9].split("E")
                R[0] = R[0][1:]
                r_itog = float(R[0]) * 10 ** int(R[1])
                self.R_itog_array.append(r_itog)
                volt_array.append(int(self.position_V))
                time_array.append(int(self.time_izm.currentText()) * 60)
                C = result_array[12].split("E")
                C[0] = C[0][1:]
                C_itog = float(C[0]) * 10 ** int(C[1])
                self.C = C_itog
                I = result_array[8].split("E")
                I[0] = I[0][1:]
                I_itog = float(I[0]) * 10 ** int(I[1])
                self.I = I_itog
                # ток в степени -8(элемент 8)
                ser.close()
                time_array.pop(1)
                time_array.pop(0)
                volt_array.pop(0)
                volt_array.pop(1)
                R_array.pop(1)
                R_array.pop(0)
                p = np.polyfit(time_array, R_array, 4)
                R_apr = np.polyval(p, time_array)
                process.kill()
                self.graphWidget.clear()
                self.graphWidget.plot(time_array, R_array, pen=pg.mkPen(color='b', width=5), name='R измеренное')
                self.graphWidget.plot(time_array, R_apr, pen=pg.mkPen(color='k', width=5), name='R апроксимированное')
                self.calculate_itog(time_array, volt_array, R_array)

                df = pd.read_csv("metadata.csv")
                df.loc[0,"number_measurment"] = df.loc[0,"number_measurment"] + 1
                df.to_csv("metadata.csv", index=False)

        except serial.SerialException as e:
            print(f"Error: {e}")


    def calculate_itog(self, Tizm, Uizm, R_meas):
        I_t = np.array(Uizm) / np.array(R_meas)

        p = np.polyfit(Tizm, R_meas, 4)
        R_apr = np.polyval(p, Tizm)
        if int(self.time_izm.currentText()) * 60 > 100:
            I_apr = np.polyval(np.polyfit(Tizm[21:], I_t[21:], 4), Tizm)

        if len(R_apr) > 12:
            DAR = R_apr[11] / R_apr[5]
        else:
            DAR = 0
        self.DAR.setText(str(round(DAR, 3)))
        if (int(self.time_izm.currentText()) * 60 // 5 + 1 < 121):
            PI = 0
            DD = 0
        else:
            PI = R_apr[117] / R_apr[9]
            DD = 1000 * (R_apr[118] - R_apr[11]) / (Uizm[1] * self.C)

        self.PI.setText(str(round(PI, 3)))
        self.DD.setText(str(round(DD, 3)))

        if PI == 0:
            W = 0
            TPI = 0
            Res = 0
        else:
            W = 3.275 - 4.819 * math.log10(PI)
            TPI = 59.029 * PI - 56.391
            Res = (70 - 45) * ((TPI / 3) ** 0.251 - 1)

        self.W.setText(str(round(W, 3)))
        self.Res.setText(str(math.trunc(Res)))

        Kabs = R_apr[10] / R_apr[1]
        self.Kabs.setText(str(round(Kabs, 3)))
        DP = 200 * TPI ** 0.251
        self.DP.setText(str(round(DP, 3)))
        R15 = R_apr[1] / 10 ** 9
        self.R15.setText(str(round(R15, 3)))
        R60 = R_apr[10] / 10 ** 9
        self.R60.setText(str(round(R60, 3)))
        R30= R_apr[4] / 10 ** 9
        self.R30.setText(str(round(R30, 3)))
        R600 = R_apr[118] / 10 ** 9
        self.R600.setText(str(round(R600, 3)))
        # if time > 100:
        #     I_ut = min(I_apr)
        #     I_spectr = (I_apr - I_ut) * time  # особое внимание этой строчке

class SettingsWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(SettingsWindow, self).__init__()
        uic.loadUi('secondUI.ui', self)

        self.keyboard = self.findChild(QtWidgets.QPushButton, 'keyboard_button')
        self.name_obj = self.findChild(QtWidgets.QTextEdit, 'name_obj')
        self.location = self.findChild(QtWidgets.QTextEdit, 'location_obj')
        self.date_field = self.findChild(QtWidgets.QDateEdit, 'dateEdit')
        self.operator = self.findChild(QtWidgets.QTextEdit, 'operator_2')
        self.save_button = self.findChild(QtWidgets.QPushButton, 'save_button')
        self.number_measurment = self.findChild(QtWidgets.QTextEdit, 'number_measurment')
        self.close_button = self.findChild(QtWidgets.QPushButton, 'close_button')

        self.keyboard.clicked.connect(self.showKeyboard)
        self.save_button.clicked.connect(self.saveSettings)
        self.close_button.clicked.connect(self.closeWindow)

        # Чтение CSV-файла при инициализации
        df = pd.read_csv('metadata.csv')

        # Установка значений в поля интерфейса
        self.name_obj.setText(str(df.loc[0, "object"]))  
        self.location.setText(str(df.loc[0, "location"]))
        self.date_field.setDate(str(df.loc[0, "date"]))
        self.operator.setText(str(df.loc[0, "operator"]))
        self.number_measurment.setText(str(df.loc[0, "number_measurment"]))  

    def closeWindow(self):
        self.close()
    def saveSettings(self):
        df = pd.read_csv('metadata.csv')
        # Обновление данных в DataFrame
        df.loc[0, "object"] = self.name_obj.toPlainText()  
        df.loc[0, "location"] = self.location.toPlainText()
        df.loc[0, "date"] = self.date_field.date().toString()
        df.loc[0, "operator"] = self.operator.toPlainText()  
        df.loc[0, "number_measurment"] = self.number_measurment.toPlainText()  
        # Сохранение изменений в CSV-файл
        df.to_csv('metadata.csv', index=False)

        subprocess.Popen(['python3', 'save_window.py'])
        #subprocess.Popen(['onboard --quit'])


    def showKeyboard(self):
        print("click button")
        subprocess.Popen(['onboard'])


class ButtonThread(QThread):
    button_pressed = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.running = True
        self.processing = False  # флаг для блокировки повторных сигналов

    def run(self):
        while self.running:
            try:
                if not GPIO.input(35):  # кнопка нажата
                    if not self.processing:
                        self.processing = True
                        self.button_pressed.emit()
                else:
                    self.processing = False  # кнопка отпущена, можно снова реагировать
            except Exception as e:
                print(f"GPIO read error in ButtonThread: {e}")
            time.sleep(0.05)  # частая проверка

    def stop(self):
        self.running = False

class GPIOMonitorThread(QThread):
    position_changed = pyqtSignal(int)

    def __init__(self):
        super().__init__()
        self.running = True

    def run(self):
        while self.running:
            try:
                if not GPIO.input(23):
                    self.position_changed.emit(500)
                elif not GPIO.input(21):
                    self.position_changed.emit(1000)
                elif not GPIO.input(19):
                    self.position_changed.emit(2500)
            except Exception as e:
                print(f"GPIO read error in thread: {e}")
            time.sleep(0.5)

    def stop(self):
        self.running = False

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    app.exec_()
